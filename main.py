from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from PIL import Image
import pytesseract
import io
import mysql.connector
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import json
import re
from groq import Groq
import dotenv
# -----------------------------
# Initialize FastAPI and Groq
# -----------------------------



dotenv.load_dotenv() 


app = FastAPI()

# Add CORS middleware to allow frontend requests
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, replace with specific origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


client = Groq(api_key=os.getenv("GROQ_API_KEY"))

# -----------------------------
# MySQL connection setup
# -----------------------------
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="3012",
    database="invoice_db"
)
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS invoices (
    id INT AUTO_INCREMENT PRIMARY KEY,
    client_name VARCHAR(255),
    invoice_number VARCHAR(100),
    invoice_date DATE,
    due_date DATE,
    subtotal DECIMAL(10,2),
    tax DECIMAL(10,2),
    total DECIMAL(10,2),
    status VARCHAR(20) DEFAULT 'Unpaid'
)
""")
conn.commit()

# -----------------------------
# Excel setup
# -----------------------------
EXCEL_FILE = "invoices.xlsx"
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Client Name", "Invoice Number", "Invoice Date", "Due Date", "Subtotal", "Tax", "Total", "Status"])
    wb.save(EXCEL_FILE)

# -----------------------------
# Helper functions
# -----------------------------
def format_date(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%Y-%m-%d")
    except:
        try:
            return datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m-%d")
        except:
            return None


def parse_invoice_text(invoice_text: str) -> dict:
    """Use Groq AI to parse invoice data and clean JSON response."""
    prompt = f"""
You are an expert invoice parsing assistant.
Extract the following fields from this invoice text and return only valid JSON:
client_name, invoice_number, invoice_date, due_date, subtotal, tax, total.

Return ONLY valid JSON — no markdown, no explanation.
Invoice text:
{invoice_text}
"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0
    )

    # ✅ Correct attribute access
    result_text = response.choices[0].message.content.strip()

    # Remove Markdown formatting if present
    result_text = re.sub(r"^```json", "", result_text)
    result_text = re.sub(r"```$", "", result_text)
    result_text = result_text.strip()

    try:
        data = json.loads(result_text)
        return data
    except json.JSONDecodeError:
        raise ValueError(f"Groq AI did not return valid JSON. Response: {result_text}")


# -----------------------------
# POST - Upload Invoice
# -----------------------------
@app.post("/upload-invoice/")
async def upload_invoice(file: UploadFile = File(...)):
    try:
        # Read image and extract text
        image_data = await file.read()
        image = Image.open(io.BytesIO(image_data))
        ocr_text = pytesseract.image_to_string(image)

        # Parse invoice using Groq AI
        invoice_data = parse_invoice_text(ocr_text)

        # Format dates
        invoice_data["invoice_date"] = format_date(invoice_data.get("invoice_date", ""))
        invoice_data["due_date"] = format_date(invoice_data.get("due_date", ""))

        # Insert into MySQL
        cursor.execute("""
            INSERT INTO invoices (client_name, invoice_number, invoice_date, due_date, subtotal, tax, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            invoice_data["client_name"],
            invoice_data["invoice_number"],
            invoice_data["invoice_date"],
            invoice_data["due_date"],
            invoice_data["subtotal"],
            invoice_data["tax"],
            invoice_data["total"]
        ))
        conn.commit()

        # Append to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            invoice_data["client_name"],
            invoice_data["invoice_number"],
            invoice_data["invoice_date"],
            invoice_data["due_date"],
            invoice_data["subtotal"],
            invoice_data["tax"],
            invoice_data["total"],
            "Unpaid"
        ])
        wb.save(EXCEL_FILE)

        return JSONResponse(content={
            "message": "Invoice processed and saved successfully!",
            "data": invoice_data
        })

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# -----------------------------
# GET - View all invoices
# -----------------------------
    @app.get("/invoices/")
def get_all_invoices():
    try:
        cursor.execute("SELECT * FROM invoices")
        rows = cursor.fetchall()

        invoices = []
        for row in rows:
            invoices.append({
                "id": row[0],
                "client_name": row[1],
                "invoice_number": row[2],
                "invoice_date": row[3].strftime("%Y-%m-%d") if row[3] else None,
                "due_date": row[4].strftime("%Y-%m-%d") if row[4] else None,
                "subtotal": float(row[5]),
                "tax": float(row[6]),
                "total": float(row[7]),
                "status": row[8]
            })

        return {"count": len(invoices), "invoices": invoices}

    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# -----------------------------
# Pydantic model for invoice updates
# -----------------------------
class InvoiceUpdate(BaseModel):
    client_name: str
    invoice_number: str
    invoice_date: str
    due_date: str
    subtotal: float
    tax: float
    total: float
    status: str


# -----------------------------
# PUT - Update an invoice
# -----------------------------
@app.put("/invoices/{invoice_id}")
def update_invoice(invoice_id: int, invoice: InvoiceUpdate):
    try:
        # Format dates
        invoice_date = format_date(invoice.invoice_date)
        due_date = format_date(invoice.due_date)
        
        # Update in MySQL
        cursor.execute("""
            UPDATE invoices 
            SET client_name = %s, invoice_number = %s, invoice_date = %s, 
                due_date = %s, subtotal = %s, tax = %s, total = %s, status = %s
            WHERE id = %s
        """, (
            invoice.client_name,
            invoice.invoice_number,
            invoice_date,
            due_date,
            invoice.subtotal,
            invoice.tax,
            invoice.total,
            invoice.status,
            invoice_id
        ))
        conn.commit()
        
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        # Update in Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Find and update the row in Excel
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[1].value == invoice.invoice_number or (row_idx - 1 == invoice_id):
                ws.cell(row=row_idx, column=1).value = invoice.client_name
                ws.cell(row=row_idx, column=2).value = invoice.invoice_number
                ws.cell(row=row_idx, column=3).value = invoice_date
                ws.cell(row=row_idx, column=4).value = due_date
                ws.cell(row=row_idx, column=5).value = invoice.subtotal
                ws.cell(row=row_idx, column=6).value = invoice.tax
                ws.cell(row=row_idx, column=7).value = invoice.total
                ws.cell(row=row_idx, column=8).value = invoice.status
                break
        
        wb.save(EXCEL_FILE)
        
        return {"message": "Invoice updated successfully", "id": invoice_id}
    
    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# -----------------------------
# DELETE - Delete an invoice
# -----------------------------
@app.delete("/invoices/{invoice_id}")
def delete_invoice(invoice_id: int):
    try:
        # Get invoice number before deleting
        cursor.execute("SELECT invoice_number FROM invoices WHERE id = %s", (invoice_id,))
        result = cursor.fetchone()
        
        if not result:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        invoice_number = result[0]
        
        # Delete from MySQL
        cursor.execute("DELETE FROM invoices WHERE id = %s", (invoice_id,))
        conn.commit()
        
        # Delete from Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Find and delete the row in Excel
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[1].value == invoice_number:
                ws.delete_rows(row_idx)
                break
        
        wb.save(EXCEL_FILE)
        
        return {"message": "Invoice deleted successfully", "id": invoice_id}
    
    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


# -----------------------------
# Serve the tracker HTML page
# -----------------------------
@app.get("/", response_class=HTMLResponse)
async def serve_tracker():
    try:
        with open("tracker.html", "r") as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>tracker.html not found</h1>", status_code=404)
